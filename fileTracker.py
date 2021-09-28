import os, shutil
from os import listdir
from openpyxl import load_workbook
import pandas as pd
from datetime import datetime
import xlrd

class fileTracker:
    def __init__(self,pas,path,filename):
        self.pas = pas
        self.path = path
        self.filename = filename
        self.currentTracker = self.getTracker()
        self.updatedTracker = self.updateTracker()
         
    
    def getTracker(self):
        df = pd.read_excel(self.path + self.filename,'Data_Ingest_Tracker')
        return df

    def updateTracker(self):
        tempTrack = self.currentTracker
        print(tempTrack['Dataset_ID'].max())
        lastID = 1 + tempTrack['Dataset_ID'].max()
        for pa in self.pas:
            for file in pa.files:
                if ~tempTrack.Filename.isin([file.name]).any() or tempTrack.loc[tempTrack['Filename']== file.name, 'Deprecated'].any(): #also check if name is in tracker but derecated
                    if len(file.tabs) > 0:
                        for tab in file.tabs:
                            tempTrack = tempTrack.append({'Dataset_ID':lastID,'PA': pa.name, 'Filepath': file.filepath, 'Filename': file.name, 'Tab_name': tab}, ignore_index=True)
                            lastID += 1
                    else:
                        tempTrack = tempTrack.append({'Dataset_ID':lastID,'PA': pa.name, 'Filepath': file.filepath, 'Filename': file.name}, ignore_index=True)
                        lastID += 1

        return tempTrack


    def exportTracker(self):
        self.archiveTracker()
        self.updatedTracker.to_excel(self.path + self.filename, sheet_name = "Sheet1", index=False)

    def archiveTracker(self):
        now = datetime.now()
        nowString = now.strftime("%Y%m%d%H%M")
        shutil.copy(self.path+self.filename, self.path + 'Archive' + '/datasetTracker' + nowString + '.xlsx')


class pa:
    def __init__(self, name, pathNames):
        self.name = name
        self.pathNames = pathNames
        self.files = self.getFiles()

    def getFiles(self):
        files = []
        for path in self.pathNames:
            dirs = listdir(path)
            fileStrs =  [x for x in dirs if os.path.isfile(os.path.join(path, x))]
            for file in fileStrs:
                filetypes = ['.csv','.txt','.xlsx','.xls', '.xlsm']
                if file.endswith(tuple(filetypes)):
                    
                    newFile = dataset(file,path)
                    files.append(newFile)
        return files
                
# never have functions with side effects - bad news bear
    
class dataset:
    def __init__(self, name, filepath):
        self.name = name
        self.filepath=filepath
        self.tabs=self.getTabs()
        

    def getTabs(self):
        tabs = []
        if self.name.endswith('.xlsx'):
            xlFile = load_workbook(os.path.join(self.filepath,self.name), read_only=True)
            for sheet in xlFile.sheetnames:
                tabs.append(sheet)
        if self.name.endswith('.xls'):
            xlFile = xlrd.open_workbook(os.path.join(self.filepath,self.name), on_demand=True)
            sheets = xlFile.sheet_names()
            for sheet in sheets:
                tabs.append(sheet)
        return tabs
    
    def updatePath(self):
        print(self.name)


PA = pa("PA NAME",["PATH TO PA DATA"])


testTracker = fileTracker([PA],"PATH TO DATASET TRACKER","DATASET TRACKER FILE NAME")
testTracker.exportTracker()
